from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Iterable, Mapping, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet


def _serialize(value):
    """Best-effort simple serialization for Excel cells."""
    if isinstance(value, (date, datetime)):
        return value
    if isinstance(value, Decimal):
        try:
            return float(value)
        except Exception:
            return str(value)
    return value


def _ensure_headers(ws: Worksheet, headers: Sequence[str]) -> None:
    """
    Ensure the first row of the worksheet contains the expected headers.
    If the sheet is empty, write headers. If headers exist but differ, raise.
    """
    if ws.max_row == 1 and ws.max_column == 1 and ws[1][0].value is None:
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)
        return

    existing = [cell.value for cell in ws[1]]
    if existing[: len(headers)] != list(headers):
        raise ValueError(
            "Excel file exists but headers do not match the expected structure."
        )


def append_row_to_xlsx(
    file_path: str | Path,
    headers: Sequence[str],
    row: Mapping[str, object] | Sequence[object] | Iterable[object],
    sheet_name: str = "Sheet1",
) -> Path:
    """
    Create or append a row to a structured XLSX file.

    Logic:
    - If the file does not exist, create it with the provided headers and first row.
    - If it exists, validate headers and append a new row using the same column order.

    Args:
        file_path: Target .xlsx file path.
        headers: Fixed column header names (defines the structure/order).
        row: Data to append. Can be one of:
             - Mapping[str, Any]: values will be ordered according to headers.
             - Sequence/Iterable: must match headers length and order.
        sheet_name: Worksheet name (default "Sheet1").

    Returns:
        Path to the written file.
    """
    path = Path(file_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    if path.exists():
        wb = load_workbook(path)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(title=sheet_name)
    else:
        wb = Workbook()
        default_ws = wb.active
        default_ws.title = sheet_name
        ws = default_ws

    _ensure_headers(ws, headers)

    if isinstance(row, Mapping):
        values = [_serialize(row.get(h)) for h in headers]
    else:
        values = list(row)
        if len(values) != len(headers):
            raise ValueError(
                "Row length does not match headers length when providing a sequence."
            )
        values = [_serialize(v) for v in values]

    ws.append(values)

    wb.save(path)
    return path


DEFAULT_HEADERS: Sequence[str] = [
    "Data użycia",
    "Godzina użycia",
    "Emerytura oczekiwana",
    "Wiek",
    "Płeć",
    "Wysokość wynagrodzenia",
    "Czy uwzględniał okresy choroby",
    "Wysokość zgromadzonych środków na koncie i Subkoncie",
    "Emerytura rzeczywista",
    "Emerytura urealniona",
    "Kod pocztowy",
]


def append_usage_row_to_xlsx(
    file_path: str | Path,
    row: Mapping[str, object] | Sequence[object] | Iterable[object],
    sheet_name: str = "Sheet1",
) -> Path:
    """
    Append a row to the default-structure usage XLSX file.

    Enforces the required fixed first-row headers (DEFAULT_HEADERS).
    Accepts a mapping keyed by those headers or a sequence in the same order.
    """
    return append_row_to_xlsx(
        file_path=file_path, headers=DEFAULT_HEADERS, row=row, sheet_name=sheet_name
    )


__all__ = ["append_row_to_xlsx", "append_usage_row_to_xlsx", "DEFAULT_HEADERS"]
